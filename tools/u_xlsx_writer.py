import openpyxl
from pathlib import Path
import string
from openpyxl.styles import (
    Alignment, Font)
from datetime import datetime


#  set columns width in sheet header
def set_column_widths(ws, columns_names, column_width):
    alphabet = string.ascii_uppercase
    for i in range(len(columns_names)):
        if type(column_width) is int:
            ws.column_dimensions[alphabet[i]].width = column_width
        elif type(column_width) is tuple:
            ws.column_dimensions[alphabet[i]].width = column_width[i]
        ws.cell(row=1, column=i + 1).font = Font(color="FF0000", size=14, bold=True)
        ws.cell(row=1, column=i + 1).alignment = Alignment(horizontal='center')


# Write the column headers if sheet is empty
def write_column_headers(ws, columns_names):
    if ws.max_row == 1:
        for col_num, column in enumerate(columns_names, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column


# Write  data from list in  the row
def write_list_in_line(ws, row_data):
    last_line = ws.max_row
    for col_num, cell_value in enumerate(row_data, 1):
        cell = ws.cell(row=last_line + 1, column=col_num)
        ws.cell(row=last_line + 1, column=col_num).alignment = Alignment(wrap_text=True, horizontal='center')
        cell.value = cell_value


# Write  data from list in  the column
def write_list_to_column(ws, column_data):
    last_column = ws.max_column
    for row_num, cell_value in enumerate(column_data, 1):
        cell = ws.cell(row=row_num, column=last_column + 1)
        ws.cell(row=row_num, column=last_column + 1).alignment = Alignment(wrap_text=True, horizontal='center')
        cell.value = cell_value


def check_sheet_name(sheet_name, workbook):
    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
    else:
        print("Worksheet not found")
        quit()
    return worksheet


def sum_column(worksheet, column_letter):
    column = worksheet[column_letter]
    return sum(cell.value for cell in column if cell.value)


def write_sum_to_selected_column(file_path: str, sheet_name:str, column_number: int):
    workbook = openpyxl.load_workbook(file_path, read_only=False)

    worksheet = check_sheet_name(sheet_name, workbook)

    last_line = worksheet.max_row

    # Get the column cells to sum
    column_cells = list(worksheet.columns)[column_number - 1]

    cell = worksheet.cell(row=last_line + 1, column=column_number)

    worksheet.cell(row=last_line + 1, column=column_number).alignment = Alignment(wrap_text=True, horizontal='center')

    cell.value = sum(int(cell.value) for cell in column_cells if cell.value is not None and cell.value != 'Sales')

    workbook.save(file_path)


def write_to_cell(ws, row_line, column_number, cell_data, photographer):
    cell = ws.cell(row=row_line + 1, column=1)
    cell.value = photographer
    cell = ws.cell(row=row_line + 1, column=column_number + 1)
    ws.cell(row=row_line + 1, column=column_number + 1).alignment = Alignment(wrap_text=True, horizontal='center')
    cell.value = cell_data


def universal_xlsx_writer(columns_names, file_path, sheet_name, photographer=None, row_line=None, column_number=None,
                          cell_data=None, row_data=None, column_width=20):
    # Create the file if it doesn't exist yet
    if not Path(file_path).is_file():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        set_column_widths(ws, columns_names, column_width)
        wb.save(file_path)

    wb = openpyxl.load_workbook(file_path, read_only=False)

    # check sheetname and create new if it doesn't exist
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.title = sheet_name
        set_column_widths(ws, columns_names, column_width)
        write_column_headers(ws, columns_names)
    ws = wb[sheet_name]
    write_column_headers(ws, columns_names)

    if cell_data is not None and column_number is not None:
        write_to_cell(ws, row_line, column_number, cell_data, photographer)
    elif row_data is not None:
        write_list_in_line(ws, row_data)

    wb.save(file_path)


if __name__ == '__main__':
    write_sum_to_selected_column('/Volumes/big4photo/Documents/Kommersant/external_sales.xlsx',
                                 '14.11.2023', 1)
